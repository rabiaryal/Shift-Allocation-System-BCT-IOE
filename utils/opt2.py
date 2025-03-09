import pandas as pd
import random
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import django
import sys
from configure.models import Employee 
from swap.models import ShiftSchedule, Shift
from django.db import transaction

sys.path.append('C:\\Users\\ASUS\\Desktop\\SASnew')

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'backend.settings')

django.setup()

# Maximum working hours for each designation
MAX_WORKING_HOURS = {
    "Cashier": 46,
    "Inventory Manager": 38,
    "Customer Help": 42,
    "Cleaning Staff": 48,
    "Manager": 58,
    "Supervisor": 56
}

# Maximum shifts an employee can work per day
MAX_SHIFTS_PER_DAY = 2


def iterate(employees):
    employees_by_designation = {}
    for emp in employees:
        designation = emp.get("designation", "Unknown") 
        if designation not in employees_by_designation:
            employees_by_designation[designation] = []
        employees_by_designation[designation].append(emp) 

    return employees_by_designation

def assign_shifts(shifts, employees_by_designation, designation_counts, MAX_WORKING_HOURS, MAX_SHIFTS_PER_DAY):
    schedule = []  # Stores shift assignments

    with transaction.atomic():  # Ensure atomic DB transactions
        for day, day_shifts in shifts.items():
            assigned_today = {}  # Track shifts per employee per day

            for shift in day_shifts:
                shift_id = shift["shift_id"]
                shift_duration = shift["shift_duration"]
                shift_timing = shift["shift_timing"]

                #  Ensure shifts are unique per day
                shift_obj, _ = Shift.objects.get_or_create(
                    shift_id=shift_id, day=day, 
                    defaults={"shift_duration": shift_duration, "shift_timing": shift_timing}
                )

                for designation, req_dict in designation_counts.items():
                    required_count = req_dict.get(shift_id, 0)

                    available_emps = [
                        emp for emp in employees_by_designation.get(designation, [])
                        if assigned_today.get(emp["e_id"], 0) < MAX_SHIFTS_PER_DAY and
                        (emp["no_of_hours_worked"] + shift_duration <= MAX_WORKING_HOURS[designation])
                    ]

                    # Sort employees by least hours worked
                    available_emps.sort(key=lambda x: x["no_of_hours_worked"])

                    for _ in range(required_count):
                        if available_emps:
                            chosen_emp = available_emps.pop(0)  # Pick employee with least hours worked
                            chosen_emp["no_of_hours_worked"] += shift_duration
                            assigned_today[chosen_emp["e_id"]] = assigned_today.get(chosen_emp["e_id"], 0) + 1

                            emp_obj = Employee.objects.get(e_id=chosen_emp["e_id"])  # Fetch employee from DB

                            # ✅ Fix ShiftSchedule creation (No 'day' field in model)
                            shift_schedule_obj, created = ShiftSchedule.objects.get_or_create(
                                shift=shift_obj, employee=emp_obj
                            )

                            print(f" Assigned {emp_obj.e_name} to {shift_obj.shift_id} on {shift_obj.day}")

                            schedule.append({
                                "Day": shift_obj.day,
                                "Shift": shift_obj.shift_id,
                                "Shift Timing": shift_obj.shift_timing,
                                "Designation": designation,
                                "Employee ID": chosen_emp["e_id"],
                                "Employee Name": chosen_emp["e_name"],
                                "Email": chosen_emp["e_gmail"],
                                "Hours Assigned": shift_duration,
                                "Total Hours After Assignment": chosen_emp["no_of_hours_worked"]
                            })
                        else:
                            print(f" No employees available for {designation} on {day}")

    return schedule








# Color dictionary for each designation
designation_colors = {
    "Cashier": "64B5F6",  # Light blue (Distinct but not too bright)
    "Inventory Manager": "FFCDD2",  # Soft pink (More distinct but gentle)
    "Customer Help": "81C784",  # Soft green (Slightly brighter but easy on the eyes)
    "Cleaning Staff": "FFEB3B",  # Warm yellow (Distinct but not overwhelming)
    "Manager": "FF9800",  # Orange (Bright, yet soft and not glaring)
    "Supervisor": "A5D6A7",  # Light green (Soft but clearer)
    "Leave": "B0BEC5"  # Grayish-blue (Subtle but clearly defined)
}

def export_schedule_to_excel(schedule, employees, output_file):
    # Prepare the schedule structure
    formatted_schedule = {
        "Employee ID": [],
        "Employee Name": [],
        "Designation": [],
        "Working Hours": [],
    }

    # Extract unique days from schedule
    unique_days = sorted(set(entry["Day"] for entry in schedule))

    # Add columns for each day
    for day in unique_days:
        formatted_schedule[day] = []

    # Dictionary to store total working hours per employee
    total_working_hours = {emp["e_id"]: emp["no_of_hours_worked"] for emp in employees}

    # Populate the formatted schedule
    for employee in employees:
        formatted_schedule["Employee ID"].append(employee["e_id"])
        formatted_schedule["Employee Name"].append(employee["e_name"])
        formatted_schedule["Designation"].append(employee["designation"])
        formatted_schedule["Working Hours"].append(total_working_hours.get(employee["e_id"], 0))

        for day in unique_days:
            assigned_shifts = [
                entry["Shift"] for entry in schedule
                if entry["Day"] == day and entry["Employee ID"] == employee["e_id"]
            ]
            formatted_schedule[day].append(
                ", ".join(assigned_shifts) if assigned_shifts else "Leave"
            )

    # Convert to DataFrame
    df = pd.DataFrame(formatted_schedule)

    # Save the dataframe to Excel
    df.to_excel(output_file, index=False)

    # Load workbook for formatting
    wb = load_workbook(output_file)
    ws = wb.active

    # Apply colors based on designation
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        designation = row[2].value  # The 'Designation' column is at index 2 (third column)
        fill_color = designation_colors.get(designation, "FFFFFF")  # Default to white if no match

        # Apply background color to the entire row
        for cell in row:
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    # Save the formatted file
    wb.save(output_file)
    print(f"Shift schedule with colors exported to {output_file}")
    os.startfile(output_file)



# def export_updated_schedule(output_dir="exports/"):
#     os.makedirs(output_dir, exist_ok=True)

#     original_file = os.path.join(output_dir, "shift_schedule_original.xlsx")
#     updated_file = os.path.join(output_dir, "shift_schedule_updated.xlsx")

#     # Fetch all shift schedules
#     schedules = ShiftSchedule.objects.select_related("shift", "employee").all()

#     # Create a structured data format
#     formatted_schedule = {
#         "Employee ID": [],
#         "Employee Name": [],
#         "Designation": [],
#         "Working Hours": [],
#     }

#     unique_days = ["Day 1", "Day 2", "Day 3", "Day 4", "Day 5", "Day 6", "Day 7"]

#     for day in unique_days:
#         formatted_schedule[day] = []

#     for schedule in schedules:
#         emp = schedule.employee
#         shift = schedule.shift

#         formatted_schedule["Employee ID"].append(emp.e_id)
#         formatted_schedule["Employee Name"].append(emp.e_name)
#         formatted_schedule["Designation"].append(emp.designation)
#         formatted_schedule["Working Hours"].append(emp.no_of_hours_worked)

#         for day in unique_days:
#             if schedule.day == day:
#                 formatted_schedule[day].append(shift.shift_id)

#     df = pd.DataFrame(formatted_schedule)

#     # Save updated schedule with a new filename
#     df.to_excel(updated_file, index=False)

#     print(f"Updated shift schedule exported to {updated_file}")
#     os.startfile(updated_file)

# def export_updated_schedule(output_file):
#     # Fetch all shift schedules
#     schedules = ShiftSchedule.objects.select_related('shift', 'employee').all()

#     # Create a structured data format
#     formatted_schedule = {
#         "Employee ID": [],
#         "Employee Name": [],
#         "Designation": [],
#         "Working Hours": [],
#     }

#     unique_days = ["Day 1", "Day 2", "Day 3", "Day 4", "Day 5", "Day 6", "Day 7"]
    
#     for day in unique_days:
#         formatted_schedule[day] = []

#     for schedule in schedules:
#         emp = schedule.employee
#         shift = schedule.shift

#         formatted_schedule["Employee ID"].append(emp.e_id)
#         formatted_schedule["Employee Name"].append(emp.e_name)
#         formatted_schedule["Designation"].append(emp.designation)
#         formatted_schedule["Working Hours"].append(emp.no_of_hours_worked)

#         for day in unique_days:
#             if schedule.day == day:
#                 formatted_schedule[day].append(shift.shift_id)

#     df = pd.DataFrame(formatted_schedule)
#     df.to_excel(output_file, index=False)

#     print(f"Updated shift schedule exported to {output_file}")
#     os.startfile(output_file)